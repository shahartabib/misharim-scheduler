# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from pyluach import dates
from datetime import date, timedelta
import calendar

# צבעים
COLORS = {
    'header': 'E6E6FA',          # סגול בהיר - כותרות
    'shabbat': 'BBDEFB',         # כחול בהיר - שבת
    'friday': 'E3F2FD',          # כחול בהיר מאוד - שישי
    'jewish_holiday': 'FFECB3', # צהוב בהיר - חגים יהודיים
    'arab_holiday': 'C8E6C9',   # ירוק בהיר - חגים ערביים
    'druze_holiday': 'D1C4E9', # סגול בהיר - חגים דרוזיים
    'vacation': 'FFE0B2',       # כתום בהיר - חופשות
    'memorial': 'FFCDD2',       # אדום בהיר - ימי זיכרון
    'month_header': '90CAF9',   # כחול - כותרת חודש
}

# חגים יהודיים תשפ"ז (5787) - מבוסס על לוח עברי
JEWISH_HOLIDAYS = {
    # תשרי תשפ"ז
    (2026, 9, 11): "ערב ר\"ה",
    (2026, 9, 12): "ראש השנה א",
    (2026, 9, 13): "ראש השנה ב",
    (2026, 9, 14): "צום גדליה",
    (2026, 9, 20): "ערב יו\"כ",
    (2026, 9, 21): "יום כיפור",
    (2026, 9, 25): "ערב סוכות",
    (2026, 9, 26): "סוכות",
    (2026, 9, 27): "חוה\"מ סוכות",
    (2026, 9, 28): "חוה\"מ סוכות",
    (2026, 9, 29): "חוה\"מ סוכות",
    (2026, 9, 30): "חוה\"מ סוכות",
    (2026, 10, 1): "חוה\"מ סוכות",
    (2026, 10, 2): "הושענא רבה",
    (2026, 10, 3): "שמיני עצרת",
    (2026, 10, 4): "שמחת תורה",
    # חנוכה
    (2026, 12, 6): "חנוכה",
    (2026, 12, 7): "חנוכה",
    (2026, 12, 8): "חנוכה",
    (2026, 12, 9): "חנוכה",
    (2026, 12, 10): "חנוכה",
    (2026, 12, 11): "חנוכה",
    (2026, 12, 12): "חנוכה",
    (2026, 12, 13): "חנוכה",
    # צומות וחגים נוספים
    (2026, 12, 24): "צום י' בטבת",
    (2027, 2, 6): "ט\"ו בשבט",
    (2027, 2, 25): "תענית אסתר",
    (2027, 2, 28): "פורים",
    (2027, 3, 1): "שושן פורים",
    # פסח
    (2027, 4, 2): "ערב פסח",
    (2027, 4, 3): "פסח",
    (2027, 4, 4): "חוה\"מ פסח",
    (2027, 4, 5): "חוה\"מ פסח",
    (2027, 4, 6): "חוה\"מ פסח",
    (2027, 4, 7): "חוה\"מ פסח",
    (2027, 4, 8): "חוה\"מ פסח",
    (2027, 4, 9): "שביעי של פסח",
    (2027, 4, 10): "אסרו חג",
    # ימי זיכרון ועצמאות
    (2027, 4, 21): "יוה\"ז לשואה",
    (2027, 4, 28): "יוה\"ז לחללים",
    (2027, 4, 29): "יום העצמאות",
    # ל"ג בעומר ושבועות
    (2027, 5, 11): "ל\"ג בעומר",
    (2027, 5, 22): "ערב שבועות",
    (2027, 5, 23): "שבועות",
    (2027, 5, 25): "יום ירושלים",
    # צומות קיץ
    (2027, 7, 4): "צום י\"ז בתמוז",
    (2027, 7, 25): "צום ט' באב",
}

# חגים ערביים (משוער - לפי לוח הירח)
ARAB_HOLIDAYS = {
    (2026, 9, 16): "ראס אל-סנה",
    (2026, 11, 25): "מולד הנביא",
    (2027, 3, 1): "תחילת הרמדאן",
    (2027, 3, 29): "ליילת אל-קדר",
    (2027, 3, 31): "עיד אל-פיטר",
    (2027, 4, 1): "עיד אל-פיטר",
    (2027, 4, 2): "עיד אל-פיטר",
    (2027, 6, 7): "עיד אל-אדחא",
    (2027, 6, 8): "עיד אל-אדחא",
    (2027, 6, 9): "עיד אל-אדחא",
    (2027, 6, 10): "עיד אל-אדחא",
}

# חגי העדה הדרוזית
DRUZE_HOLIDAYS = {
    (2027, 4, 25): "חג הנביא שועייב",
    (2027, 8, 1): "חג הנביא סבלאן",
}

# ימי זיכרון
MEMORIAL_DAYS = {
    (2026, 10, 7): "יוה\"ז ל-7.10",
    (2026, 11, 4): "יוה\"ז ליצחק רבין",
}

def get_hebrew_date_str(d):
    """מחזיר תאריך עברי בפורמט קריא"""
    try:
        heb = dates.HebrewDate.from_pydate(d)
        day = heb.day
        # המרה לאותיות עבריות
        hebrew_nums = {
            1: "א'", 2: "ב'", 3: "ג'", 4: "ד'", 5: "ה'", 6: "ו'", 7: "ז'", 8: "ח'", 9: "ט'",
            10: "י'", 11: 'י"א', 12: 'י"ב', 13: 'י"ג', 14: 'י"ד', 15: 'ט"ו', 16: 'ט"ז',
            17: 'י"ז', 18: 'י"ח', 19: 'י"ט', 20: "כ'", 21: 'כ"א', 22: 'כ"ב', 23: 'כ"ג',
            24: 'כ"ד', 25: 'כ"ה', 26: 'כ"ו', 27: 'כ"ז', 28: 'כ"ח', 29: 'כ"ט', 30: "ל'"
        }
        return hebrew_nums.get(day, str(day))
    except:
        return ""

def get_day_of_week_hebrew(d):
    """מחזיר יום בשבוע בעברית"""
    days = ["ב'", "ג'", "ד'", "ה'", "ו'", "ש'", "א'"]
    return days[d.weekday()]

# יצירת workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "גאנט תשפז"
ws.sheet_view.rightToLeft = True

# הגדרות גופן וגבול
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# כותרת ראשית
ws.merge_cells('A1:AK1')
ws['A1'] = 'לוח שנתי מישרים תשפ"ז (אוגוסט 2026 - אוגוסט 2027)'
ws['A1'].font = Font(size=18, bold=True)
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws['A1'].fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type='solid')

# מקרא צבעים
legend_row = 2
ws.merge_cells('A2:B2')
ws['A2'] = 'מקרא:'
ws['A2'].font = Font(bold=True)

legend_items = [
    ('C2', 'שבת', COLORS['shabbat']),
    ('D2', 'שישי', COLORS['friday']),
    ('E2', 'חג יהודי', COLORS['jewish_holiday']),
    ('F2', 'חג ערבי', COLORS['arab_holiday']),
    ('G2', 'חג דרוזי', COLORS['druze_holiday']),
    ('H2', 'חופשה', COLORS['vacation']),
    ('I2', 'יום זיכרון', COLORS['memorial']),
]

for cell, text, color in legend_items:
    ws[cell] = text
    ws[cell].fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    ws[cell].alignment = Alignment(horizontal='center')
    ws[cell].border = thin_border

# החודשים לתשפ"ז - מאוגוסט 2026 עד אוגוסט 2027
months = [
    (2026, 8, 'אוגוסט 2026'),
    (2026, 9, 'ספטמבר 2026'),
    (2026, 10, 'אוקטובר 2026'),
    (2026, 11, 'נובמבר 2026'),
    (2026, 12, 'דצמבר 2026'),
    (2027, 1, 'ינואר 2027'),
    (2027, 2, 'פברואר 2027'),
    (2027, 3, 'מרץ 2027'),
    (2027, 4, 'אפריל 2027'),
    (2027, 5, 'מאי 2027'),
    (2027, 6, 'יוני 2027'),
    (2027, 7, 'יולי 2027'),
    (2027, 8, 'אוגוסט 2027'),
]

current_row = 4

for year, month, month_name in months:
    # כותרת חודש
    ws.merge_cells(f'A{current_row}:AK{current_row}')
    ws[f'A{current_row}'] = month_name
    ws[f'A{current_row}'].font = Font(size=14, bold=True, color='FFFFFF')
    ws[f'A{current_row}'].fill = PatternFill(start_color=COLORS['month_header'], end_color=COLORS['month_header'], fill_type='solid')
    ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
    current_row += 1

    # שורת ימים בשבוע
    header_row = current_row
    ws[f'A{header_row}'] = 'יום'
    ws[f'A{header_row}'].font = Font(bold=True)
    ws[f'A{header_row}'].border = thin_border
    ws[f'A{header_row}'].alignment = Alignment(horizontal='center')

    # קבלת מספר ימים בחודש
    num_days = calendar.monthrange(year, month)[1]

    for day in range(1, num_days + 1):
        col = day + 1  # עמודה B היא יום 1
        col_letter = get_column_letter(col)
        d = date(year, month, day)
        ws[f'{col_letter}{header_row}'] = get_day_of_week_hebrew(d)
        ws[f'{col_letter}{header_row}'].font = Font(bold=True, size=9)
        ws[f'{col_letter}{header_row}'].alignment = Alignment(horizontal='center')
        ws[f'{col_letter}{header_row}'].border = thin_border

        # צבע לשישי/שבת
        if d.weekday() == 5:  # שבת
            ws[f'{col_letter}{header_row}'].fill = PatternFill(start_color=COLORS['shabbat'], end_color=COLORS['shabbat'], fill_type='solid')
        elif d.weekday() == 4:  # שישי
            ws[f'{col_letter}{header_row}'].fill = PatternFill(start_color=COLORS['friday'], end_color=COLORS['friday'], fill_type='solid')

    current_row += 1

    # שורת תאריך לועזי
    greg_row = current_row
    ws[f'A{greg_row}'] = 'לועזי'
    ws[f'A{greg_row}'].font = Font(bold=True, size=9)
    ws[f'A{greg_row}'].border = thin_border
    ws[f'A{greg_row}'].alignment = Alignment(horizontal='center')

    for day in range(1, num_days + 1):
        col = day + 1
        col_letter = get_column_letter(col)
        d = date(year, month, day)
        ws[f'{col_letter}{greg_row}'] = day
        ws[f'{col_letter}{greg_row}'].alignment = Alignment(horizontal='center')
        ws[f'{col_letter}{greg_row}'].border = thin_border
        ws[f'{col_letter}{greg_row}'].font = Font(size=10)

        # צבע לשישי/שבת
        if d.weekday() == 5:
            ws[f'{col_letter}{greg_row}'].fill = PatternFill(start_color=COLORS['shabbat'], end_color=COLORS['shabbat'], fill_type='solid')
        elif d.weekday() == 4:
            ws[f'{col_letter}{greg_row}'].fill = PatternFill(start_color=COLORS['friday'], end_color=COLORS['friday'], fill_type='solid')

    current_row += 1

    # שורת תאריך עברי
    heb_row = current_row
    ws[f'A{heb_row}'] = 'עברי'
    ws[f'A{heb_row}'].font = Font(bold=True, size=9)
    ws[f'A{heb_row}'].border = thin_border
    ws[f'A{heb_row}'].alignment = Alignment(horizontal='center')

    for day in range(1, num_days + 1):
        col = day + 1
        col_letter = get_column_letter(col)
        d = date(year, month, day)
        ws[f'{col_letter}{heb_row}'] = get_hebrew_date_str(d)
        ws[f'{col_letter}{heb_row}'].alignment = Alignment(horizontal='center')
        ws[f'{col_letter}{heb_row}'].border = thin_border
        ws[f'{col_letter}{heb_row}'].font = Font(size=9)

        # צבע לשישי/שבת
        if d.weekday() == 5:
            ws[f'{col_letter}{heb_row}'].fill = PatternFill(start_color=COLORS['shabbat'], end_color=COLORS['shabbat'], fill_type='solid')
        elif d.weekday() == 4:
            ws[f'{col_letter}{heb_row}'].fill = PatternFill(start_color=COLORS['friday'], end_color=COLORS['friday'], fill_type='solid')

    current_row += 1

    # שורת חגים/אירועים
    events_row = current_row
    ws[f'A{events_row}'] = 'אירועים'
    ws[f'A{events_row}'].font = Font(bold=True, size=9)
    ws[f'A{events_row}'].border = thin_border
    ws[f'A{events_row}'].alignment = Alignment(horizontal='center')

    for day in range(1, num_days + 1):
        col = day + 1
        col_letter = get_column_letter(col)
        d = date(year, month, day)
        date_key = (year, month, day)

        event_text = ""
        fill_color = None

        # בדיקת חגים
        if date_key in JEWISH_HOLIDAYS:
            event_text = JEWISH_HOLIDAYS[date_key]
            fill_color = COLORS['jewish_holiday']
        elif date_key in ARAB_HOLIDAYS:
            event_text = ARAB_HOLIDAYS[date_key]
            fill_color = COLORS['arab_holiday']
        elif date_key in DRUZE_HOLIDAYS:
            event_text = DRUZE_HOLIDAYS[date_key]
            fill_color = COLORS['druze_holiday']
        elif date_key in MEMORIAL_DAYS:
            event_text = MEMORIAL_DAYS[date_key]
            fill_color = COLORS['memorial']
        elif d.weekday() == 5:  # שבת
            fill_color = COLORS['shabbat']
        elif d.weekday() == 4:  # שישי
            fill_color = COLORS['friday']

        ws[f'{col_letter}{events_row}'] = event_text
        ws[f'{col_letter}{events_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws[f'{col_letter}{events_row}'].border = thin_border
        ws[f'{col_letter}{events_row}'].font = Font(size=8)

        if fill_color:
            ws[f'{col_letter}{events_row}'].fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')

    current_row += 1

    # שורה ריקה לאירועים של מישרים
    misharim_row = current_row
    ws[f'A{misharim_row}'] = 'מישרים'
    ws[f'A{misharim_row}'].font = Font(bold=True, size=9)
    ws[f'A{misharim_row}'].border = thin_border
    ws[f'A{misharim_row}'].alignment = Alignment(horizontal='center')

    for day in range(1, num_days + 1):
        col = day + 1
        col_letter = get_column_letter(col)
        d = date(year, month, day)
        ws[f'{col_letter}{misharim_row}'].border = thin_border

        # צבע לשישי/שבת
        if d.weekday() == 5:
            ws[f'{col_letter}{misharim_row}'].fill = PatternFill(start_color=COLORS['shabbat'], end_color=COLORS['shabbat'], fill_type='solid')
        elif d.weekday() == 4:
            ws[f'{col_letter}{misharim_row}'].fill = PatternFill(start_color=COLORS['friday'], end_color=COLORS['friday'], fill_type='solid')

    current_row += 2  # שורה ריקה בין החודשים

# התאמת רוחב עמודות
ws.column_dimensions['A'].width = 8
for i in range(2, 33):
    ws.column_dimensions[get_column_letter(i)].width = 6

# שמירת הקובץ
output_path = r'C:\Users\Shahar Tabib\OneDrive - Sliceknowledge.com\Slice\שתפים\מישרים\רות זהר גאנט תשפז\לוח שנתי מישרים תשפז - חדש.xlsx'
wb.save(output_path)
print(f'הקובץ נשמר בהצלחה: {output_path}')
