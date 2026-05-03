# -*- coding: utf-8 -*-
"""
עדכון קובץ גאנט לתשפ"ז
מעתיק את הקובץ המקורי ומעדכן תאריכים וחגים
כולל אוגוסט 2026 עד אוגוסט 2027
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pyluach import dates
from datetime import date, timedelta
import calendar
import shutil

# קבצים
source_path = r'C:\Users\Shahar Tabib\OneDrive - Sliceknowledge.com\Slice\שתפים\מישרים\רות זהר גאנט תשפז\לוח שנתי מישרים תשפז.xlsx'
output_path = r'C:\Users\Shahar Tabib\OneDrive - Sliceknowledge.com\Slice\שתפים\מישרים\רות זהר גאנט תשפז\גאנט מישרים תשפז.xlsx'

# העתקת הקובץ המקורי
shutil.copy(source_path, output_path)

# טעינת הקובץ המועתק
wb = openpyxl.load_workbook(output_path)
ws = wb['גאנט תשפו']

# שינוי שם הגיליון
ws.title = "גאנט תשפז"


def get_hebrew_date_str(d):
    """מחזיר תאריך עברי בפורמט קריא"""
    try:
        heb = dates.HebrewDate.from_pydate(d)
        day = heb.day
        hebrew_nums = {
            1: "א'", 2: "ב'", 3: "ג'", 4: "ד'", 5: "ה'", 6: "ו'", 7: "ז'", 8: "ח'", 9: "ט'",
            10: "י'", 11: 'י"א', 12: 'י"ב', 13: 'י"ג', 14: 'י"ד', 15: 'ט"ו', 16: 'ט"ז',
            17: 'י"ז', 18: 'י"ח', 19: 'י"ט', 20: "כ'", 21: 'כ"א', 22: 'כ"ב', 23: 'כ"ג',
            24: 'כ"ד', 25: 'כ"ה', 26: 'כ"ו', 27: 'כ"ז', 28: 'כ"ח', 29: 'כ"ט', 30: "ל'"
        }
        return hebrew_nums.get(day, str(day))
    except:
        return ""


def get_weekday_hebrew(d):
    """מחזיר יום בשבוע בעברית"""
    days = ["ב'", "ג'", "ד'", "ה'", "ו'", "ש'", "א'"]
    return days[d.weekday()]


def is_merged_cell(ws, cell_coord):
    """בדיקה אם תא הוא חלק מתאים ממוזגים"""
    for merged_range in ws.merged_cells.ranges:
        if cell_coord in merged_range:
            return True
    return False


def safe_set_value(ws, cell_coord, value):
    """הגדרת ערך בתא בצורה בטוחה"""
    if not is_merged_cell(ws, cell_coord):
        ws[cell_coord] = value
        return True
    return False


# חגים ואירועים לתשפ"ז (ספטמבר 2026 - אוגוסט 2027)
# תאריכי החגים מבוססים על לוח שנה עברי
EVENTS = {
    # ספטמבר 2026
    (2026, 9, 11): 'ערה"ש',
    (2026, 9, 12): 'ראש השנה',
    (2026, 9, 13): 'ראש השנה',
    (2026, 9, 14): 'צום גדליה',
    (2026, 9, 20): 'עיו"כ',
    (2026, 9, 21): 'יום כיפור',
    (2026, 9, 22): 'חופש',
    (2026, 9, 25): 'ערב סוכות',
    (2026, 9, 26): 'סוכות',
    (2026, 9, 27): 'חוה"מ סוכות',
    (2026, 9, 28): 'חוה"מ סוכות',
    (2026, 9, 29): 'חוה"מ סוכות',
    (2026, 9, 30): 'חוה"מ סוכות',
    # אוקטובר 2026
    (2026, 10, 1): 'חוה"מ',
    (2026, 10, 2): 'הושענא רבה',
    (2026, 10, 3): 'שמחת תורה',
    (2026, 10, 7): 'יוה"ז לאירועי ה-7 באוקטובר',
    # נובמבר 2026
    (2026, 11, 4): 'יוה"ז לי.רבין',
    # דצמבר 2026 - חנוכה
    (2026, 12, 6): 'חנוכה',
    (2026, 12, 7): 'חנוכה',
    (2026, 12, 8): 'חנוכה',
    (2026, 12, 9): 'חנוכה',
    (2026, 12, 10): 'חנוכה',
    (2026, 12, 11): 'חנוכה',
    (2026, 12, 12): 'חנוכה',
    (2026, 12, 13): 'חנוכה',
    (2026, 12, 24): 'צום עשרה בטבת',
    # פברואר 2027
    (2027, 2, 6): 'ט"ו בשבט',
    (2027, 2, 25): 'ת. אסתר',
    (2027, 2, 28): 'פורים',
    # מרץ 2027
    (2027, 3, 1): 'התחלת הרמדאן',
    (2027, 3, 16): 'שאלון מורים',
    (2027, 3, 29): 'עיד אל פיטר',
    (2027, 3, 31): 'חופשת פסח בבתי ספר',
    # אפריל 2027
    (2027, 4, 2): 'ערב פסח',
    (2027, 4, 3): 'פסח',
    (2027, 4, 4): 'חוה"מ פסח',
    (2027, 4, 5): 'חוה"מ פסח',
    (2027, 4, 6): 'חוה"מ פסח',
    (2027, 4, 7): 'חוה"מ פסח',
    (2027, 4, 8): 'חוה"מ פסח',
    (2027, 4, 9): 'שביעי של פסח',
    (2027, 4, 10): 'אסרו חג',
    (2027, 4, 19): 'מחוון אוריינות',
    (2027, 4, 21): 'יוה"ז לחללי מערכות ישראל',
    (2027, 4, 22): 'יום העצמאות',
    (2027, 4, 25): 'חג הנביא שועייב',
    # מאי 2027
    (2027, 5, 5): 'ל"ג בעומר',
    (2027, 5, 15): 'יום ירושלים',
    (2027, 5, 22): 'ערב שבועות',
    (2027, 5, 23): 'שבועות',
    # יוני 2027
    (2027, 6, 7): 'עיד אל אדחא',
    (2027, 6, 8): 'עיד אל אדחא',
    (2027, 6, 9): 'עיד אל אדחא',
    (2027, 6, 30): 'משימות סוף שנה כיתות א',
    # יולי 2027
    (2027, 7, 4): 'צום י"ז בתמוז',
    (2027, 7, 25): 'צום ט\' באב',
}


# מבנה החודשים בקובץ המקורי
# שורות: ימים(5), תאריך לועזי, תאריך עברי, אירועים, שורות נוספות
# רווח של 5 שורות בין חודשים

# מציאת מבנה החודשים - שורות תאריכים לועזיים
# מהמידע שאספנו קודם:
# שורה 6: ספטמבר (1 בעמודה I)
# שורה 11: אוקטובר (1 בעמודה L)
# שורה 16: נובמבר (1 בעמודה G)
# וכו'

# מיפוי החודשים החדשים - נשתמש באותו מבנה שורות
# אבל נעדכן את התוכן לתשפ"ז

# שמות החודשים בעברית
MONTH_NAMES = {
    (2026, 8): 'אוגוסט',
    (2026, 9): 'ספטמבר',
    (2026, 10): 'אוקטובר',
    (2026, 11): 'נובמבר',
    (2026, 12): 'דצמבר',
    (2027, 1): 'ינואר',
    (2027, 2): 'פברואר',
    (2027, 3): 'מרץ',
    (2027, 4): 'אפריל',
    (2027, 5): 'מאי',
    (2027, 6): 'יוני',
    (2027, 7): 'יולי',
    (2027, 8): 'אוגוסט',
}

MONTH_CONFIG = [
    # (year, month, greg_row, heb_row, events_row, month_name_row)
    # הסדר לפי שנת הלימודים - מתחיל מאוגוסט
    # כל חודש תופס 5 שורות: יום, תאריך לועזי, תאריך עברי, אירועים, ריק
    (2026, 8, 6, 7, 8, 7),      # אוגוסט 2026 - שורות 5-9
    (2026, 9, 11, 12, 13, 12),  # ספטמבר - שורות 10-14
    (2026, 10, 16, 17, 18, 17), # אוקטובר - שורות 15-19
    (2026, 11, 21, 22, 23, 22), # נובמבר - שורות 20-24
    (2026, 12, 26, 27, 28, 27), # דצמבר - שורות 25-29
    (2027, 1, 31, 32, 33, 32),  # ינואר - שורות 30-34
    (2027, 2, 36, 37, 38, 37),  # פברואר - שורות 35-39
    (2027, 3, 41, 42, 43, 42),  # מרץ - שורות 40-44
    (2027, 4, 46, 47, 48, 47),  # אפריל - שורות 45-49
    (2027, 5, 51, 52, 53, 52),  # מאי - שורות 50-54
    (2027, 6, 56, 57, 58, 57),  # יוני - שורות 55-59
    (2027, 7, 61, 62, 63, 62),  # יולי - שורות 60-64
    (2027, 8, 66, 67, 68, 67),  # אוגוסט 2027 - שורות 65-69
]

# מסגרת דקה לתאים
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# עמודות הימים בשורה 5:
# F=ד'(Wed), G=ה'(Thu), H=ו'(Fri), I=ש'(Sat), J=א'(Sun), K=ב'(Mon), L=ג'(Tue), M=ד'...
# weekday: Mon=0, Tue=1, Wed=2, Thu=3, Fri=4, Sat=5, Sun=6

# מיפוי יום בשבוע לעמודה (כאשר ד' הוא עמודה F)
def get_column_for_date(d):
    """מחזיר את העמודה עבור תאריך מסוים"""
    # F=6 היא יום ד' (Wednesday=2)
    base_col = 6  # F
    base_weekday = 2  # Wednesday

    day_of_month = d.day
    first_of_month = date(d.year, d.month, 1)
    first_weekday = first_of_month.weekday()

    # חישוב העמודה של יום 1 בחודש
    if first_weekday >= base_weekday:
        offset = first_weekday - base_weekday
    else:
        offset = 7 - (base_weekday - first_weekday)

    first_col = base_col + offset

    # העמודה של התאריך הנוכחי
    return first_col + day_of_month - 1


# ראשית, נבטל את כל התאים הממוזגים
print("מבטל תאים ממוזגים...")
merged_to_remove = []
for merged_range in ws.merged_cells.ranges:
    # בדיקה אם ה-merge נמצא בשורות רלוונטיות או בעמודות C-E (שמות חודשים)
    event_rows = [8, 13, 18, 19, 20, 23, 28, 29, 33, 34, 38, 39, 43, 44, 48, 49, 50, 54, 55, 56, 59, 60, 61, 64, 65, 66, 67]
    # גם תאים ממוזגים בעמודות C-E
    if merged_range.min_row in event_rows or merged_range.min_col <= 5:
        merged_to_remove.append(str(merged_range))

for mr in merged_to_remove:
    try:
        ws.unmerge_cells(mr)
        print(f"  Unmerged: {mr}")
    except:
        pass

# ניקוי שורות התאריכים והאירועים ומילוי מחדש
for year, month, greg_row, heb_row, events_row, month_name_row in MONTH_CONFIG:
    num_days = calendar.monthrange(year, month)[1]

    print(f"\n{year}/{month}:")

    # ניקוי כל השורות הקיימות (עמודות F עד AP)
    for col_num in range(6, 43):
        col_letter = get_column_letter(col_num)
        # ניקוי כל השורות הרלוונטיות
        for row in [greg_row, heb_row, events_row, events_row + 1, events_row + 2]:
            cell_coord = f'{col_letter}{row}'
            try:
                ws[cell_coord] = None
            except:
                pass

    # הוספת שם החודש בעמודות C-E (ממוזגות)
    month_name = MONTH_NAMES.get((year, month), '')
    # ניסיון להוסיף שם חודש בתא C
    try:
        ws[f'C{month_name_row}'] = month_name
        ws[f'C{month_name_row}'].font = Font(bold=True, size=11)
        ws[f'C{month_name_row}'].alignment = Alignment(horizontal='center', vertical='center', text_rotation=90)
    except:
        pass

    # מילוי התאריכים החדשים
    for day in range(1, num_days + 1):
        current_date = date(year, month, day)
        col_num = get_column_for_date(current_date)
        col_letter = get_column_letter(col_num)

        if day == 1:
            print(f"  Day 1 at column {col_letter} ({col_num})")

        # תאריך לועזי
        cell_coord = f'{col_letter}{greg_row}'
        if not is_merged_cell(ws, cell_coord):
            ws[cell_coord] = day
            ws[cell_coord].border = thin_border
            ws[cell_coord].alignment = Alignment(horizontal='center', vertical='center')

        # תאריך עברי
        cell_coord = f'{col_letter}{heb_row}'
        if not is_merged_cell(ws, cell_coord):
            ws[cell_coord] = get_hebrew_date_str(current_date)
            ws[cell_coord].border = thin_border
            ws[cell_coord].alignment = Alignment(horizontal='center', vertical='center')

        # אירועים
        date_key = (year, month, day)
        cell_coord = f'{col_letter}{events_row}'
        if not is_merged_cell(ws, cell_coord):
            if date_key in EVENTS:
                ws[cell_coord] = EVENTS[date_key]
            ws[cell_coord].border = thin_border
            ws[cell_coord].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# הוספת מסגרות לכל התאים בטווח הגאנט
print("\nמוסיף מסגרות לכל התאים...")
for row in range(5, 70):
    for col_num in range(6, 43):
        col_letter = get_column_letter(col_num)
        cell_coord = f'{col_letter}{row}'
        if not is_merged_cell(ws, cell_coord):
            try:
                ws[cell_coord].border = thin_border
            except:
                pass

# שמירה
wb.save(output_path)
print(f'\nהקובץ נשמר: {output_path}')
