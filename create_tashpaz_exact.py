# -*- coding: utf-8 -*-
"""
יצירת גאנט תשפ"ז - העתקה מדויקת של המבנה המקורי
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pyluach import dates
from datetime import date
import calendar
import shutil

# קבצים
source_path = r'C:\Users\Shahar Tabib\OneDrive - Sliceknowledge.com\Slice\שתפים\מישרים\רות זהר גאנט תשפז\לוח שנתי מישרים תשפז.xlsx'
output_path = r'C:\Users\Shahar Tabib\OneDrive - Sliceknowledge.com\Slice\שתפים\מישרים\רות זהר גאנט תשפז\גאנט מישרים תשפז.xlsx'

# העתקת הקובץ המקורי
shutil.copy(source_path, output_path)

# טעינת הקובץ
wb = openpyxl.load_workbook(output_path)
ws = wb['גאנט תשפו']
ws.title = "גאנט תשפז"

# צבעים מהקובץ המקורי
SHABBAT_COLOR = 'FFE3C9B7'  # חום בהיר - שבת
FRIDAY_COLOR = 'FF99CCFF'   # כחול - שישי
WHITE_COLOR = 'FFFFFFFF'    # לבן
GREEN_COLOR = 'FFECF9E7'    # ירוק בהיר - עמודת חודשים

# מסגרת
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def get_hebrew_date_str(d):
    """מחזיר תאריך עברי בפורמט קריא"""
    try:
        heb = dates.HebrewDate.from_pydate(d)
        day = heb.day
        hebrew_nums = {
            1: "א'", 2: "ב'", 3: "ג'", 4: "ד'", 5: "ה'", 6: "ו'", 7: "ז'", 8: "ח'", 9: "ט'",
            10: "י'", 11: 'י"א', 12: 'י"ב', 13: 'י"ג', 14: 'י"ד', 15: 'ט"ו', 16: 'ט"ז',
            17: 'י"ז', 18: 'י"ח', 19: 'י"ט', 20: "כ'", 21: 'כ"א', 22: 'כ"ב', 23: 'כ"ג',
            24: 'כ"ד', 25: 'כ"ה', 26: 'כ"ו', 27: 'כ"ז', 28: 'כ"ח', 29: 'כ"ט', 30: "ל'"
        }
        return hebrew_nums.get(day, str(day))
    except:
        return ""

def is_merged_cell(ws, cell_coord):
    """בדיקה אם תא ממוזג"""
    for merged_range in ws.merged_cells.ranges:
        if cell_coord in merged_range:
            return True
    return False

def unmerge_all_events(ws):
    """ביטול כל התאים הממוזגים בשורות האירועים"""
    to_remove = []
    for merged_range in ws.merged_cells.ranges:
        # לא לגעת בתאים ממוזגים בעמודות A-E (שמות חודשים)
        if merged_range.min_col > 5:
            to_remove.append(str(merged_range))

    for mr in to_remove:
        try:
            ws.unmerge_cells(mr)
        except:
            pass

# ביטול תאים ממוזגים
print("מבטל תאים ממוזגים...")
unmerge_all_events(ws)

# מבנה החודשים בקובץ המקורי (לפי הניתוח):
# שורה 5: ימי שבוע (קבועים - F עד AO)
#
# ספטמבר: שורה 6 (לועזי), 7 (עברי), 8-10 (אירועים) - 1 מתחיל בעמודה I
# אוקטובר: שורה 11 (לועזי), 12 (עברי), 13-15 (אירועים) - 1 מתחיל בעמודה L
# נובמבר: שורה 16 (לועזי), 17 (עברי), 18-20 (אירועים) - 1 מתחיל בעמודה G
# דצמבר: שורה 21 (לועזי), 22 (עברי), 23-25 (אירועים) - 1 מתחיל בעמודה J
# ינואר: שורה 26 (לועזי), 27 (עברי), 28-30 (אירועים) - 1 מתחיל בעמודה L
# פברואר: שורה 31 (לועזי), 32 (עברי), 33-35 (אירועים) - 1 מתחיל בעמודה H
# מרץ: שורה 36 (לועזי), 37 (עברי), 38-40 (אירועים) - 1 מתחיל בעמודה K
# אפריל: שורה 41 (לועזי), 42 (עברי), 43-45 (אירועים) - 1 מתחיל בעמודה J
# מאי: שורה 46 (לועזי), 47 (עברי), 48-51 (אירועים) - 1 מתחיל בעמודה F
# יוני: שורה 52 (לועזי), 53 (עברי), 54-56 (אירועים) - 1 מתחיל בעמודה H
# יולי: שורה 57 (לועזי), 58 (עברי), 59-61 (אירועים) - 1 מתחיל בעמודה K
# אוגוסט: שורה 62 (לועזי), 63 (עברי), 64-67 (אירועים) - 1 מתחיל בעמודה F
# [שורה 68-69 - כנראה ספטמבר הבא]

# מיפוי החודשים החדשים לתשפ"ז
# תשפ"ז: ספטמבר 2026 - אוגוסט 2027
# נשמור על אותו מבנה שורות!

MONTHS_TASHPAZ = [
    # (year, month, greg_row, heb_row, events_row, month_name)
    (2026, 9, 6, 7, 8, 'ספטמבר'),
    (2026, 10, 11, 12, 13, 'אוקטובר'),
    (2026, 11, 16, 17, 18, 'נובמבר'),
    (2026, 12, 21, 22, 23, 'דצמבר'),
    (2027, 1, 26, 27, 28, 'ינואר'),
    (2027, 2, 31, 32, 33, 'פברואר'),
    (2027, 3, 36, 37, 38, 'מרץ'),
    (2027, 4, 41, 42, 43, 'אפריל'),
    (2027, 5, 46, 47, 48, 'מאי'),
    (2027, 6, 52, 53, 54, 'יוני'),
    (2027, 7, 57, 58, 59, 'יולי'),
    (2027, 8, 62, 63, 64, 'אוגוסט'),
]

# חגים יהודיים תשפ"ז
JEWISH_HOLIDAYS = {
    # ספטמבר 2026 - תשרי תשפ"ז
    (2026, 9, 11): 'ערה"ש',
    (2026, 9, 12): 'ראש השנה',
    (2026, 9, 13): 'ראש השנה',
    (2026, 9, 14): 'צום גדליה',
    (2026, 9, 20): 'עיו"כ',
    (2026, 9, 21): 'יום כיפור',
    (2026, 9, 25): 'ערב סוכות',
    (2026, 9, 26): 'סוכות',
    (2026, 9, 27): 'חוה"מ סוכות',
    (2026, 9, 28): 'חוה"מ סוכות',
    (2026, 9, 29): 'חוה"מ סוכות',
    (2026, 9, 30): 'חוה"מ סוכות',
    # אוקטובר 2026
    (2026, 10, 1): 'חוה"מ סוכות',
    (2026, 10, 2): 'הושענא רבה',
    (2026, 10, 3): 'שמחת תורה',
    (2026, 10, 7): 'יוה"ז לאירועי ה-7 באוקטובר',
    # נובמבר 2026
    (2026, 11, 4): 'יוה"ז לי.רבין',
    # דצמבר 2026 - חנוכה
    (2026, 12, 6): 'חנוכה',
    (2026, 12, 7): 'חנוכה',
    (2026, 12, 8): 'חנוכה',
    (2026, 12, 9): 'חנוכה',
    (2026, 12, 10): 'חנוכה',
    (2026, 12, 11): 'חנוכה',
    (2026, 12, 12): 'חנוכה',
    (2026, 12, 13): 'חנוכה',
    (2026, 12, 24): 'צום עשרה בטבת',
    # פברואר 2027
    (2027, 2, 6): 'ט"ו בשבט',
    (2027, 2, 25): 'ת. אסתר',
    (2027, 2, 28): 'פורים',
    # אפריל 2027 - פסח
    (2027, 4, 2): 'ערב פסח',
    (2027, 4, 3): 'פסח',
    (2027, 4, 4): 'חוה"מ פסח',
    (2027, 4, 5): 'חוה"מ פסח',
    (2027, 4, 6): 'חוה"מ פסח',
    (2027, 4, 7): 'חוה"מ פסח',
    (2027, 4, 8): 'חוה"מ פסח',
    (2027, 4, 9): 'שביעי של פסח',
    (2027, 4, 10): 'אסרו חג',
    (2027, 4, 21): 'יוה"ז לשואה ולגבורה',
    (2027, 4, 28): 'יוה"ז לחללי מערכות ישראל',
    (2027, 4, 29): 'יום העצמאות',
    # מאי 2027
    (2027, 5, 11): 'ל"ג בעומר',
    (2027, 5, 22): 'ערב שבועות',
    (2027, 5, 23): 'שבועות',
    (2027, 5, 25): 'יום ירושלים',
    # יולי 2027
    (2027, 7, 4): 'צום י"ז בתמוז',
    (2027, 7, 25): 'צום ט\' באב',
}

# חגים ערביים תשפ"ז (תאריכים משוערים לפי לוח הירח)
ARAB_HOLIDAYS = {
    (2027, 3, 1): 'התחלת הרמדאן',
    (2027, 3, 31): 'עיד אל-פיטר',
    (2027, 6, 7): 'עיד אל-אדחא',
    (2027, 6, 8): 'עיד אל-אדחא',
    (2027, 6, 9): 'עיד אל-אדחא',
    (2027, 6, 10): 'עיד אל-אדחא',
}

# חגים דרוזיים
DRUZE_HOLIDAYS = {
    (2027, 4, 25): 'חג הנביא שועייב',
}

def get_column_for_weekday(weekday, base_col=6):
    """
    מחזיר את העמודה עבור יום בשבוע מסוים
    weekday: 0=Monday, 1=Tuesday, 2=Wednesday, 3=Thursday, 4=Friday, 5=Saturday, 6=Sunday
    base_col: העמודה הראשונה (F=6 היא יום ד' = Wednesday = 2)
    """
    # עמודה F (6) = יום ד' (Wednesday = weekday 2)
    # מחזור של 7 ימים
    base_weekday = 2  # Wednesday

    if weekday >= base_weekday:
        offset = weekday - base_weekday
    else:
        offset = 7 - (base_weekday - weekday)

    return base_col + offset

def get_first_col_for_month(year, month):
    """מחזיר את העמודה של יום 1 בחודש"""
    first_day = date(year, month, 1)
    first_weekday = first_day.weekday()
    return get_column_for_weekday(first_weekday)

# ניקוי כל השורות ומילוי מחדש
print("\nמעדכן תאריכים...")

for year, month, greg_row, heb_row, events_row, month_name in MONTHS_TASHPAZ:
    print(f"\n{month_name} {year}:")

    num_days = calendar.monthrange(year, month)[1]
    first_col = get_first_col_for_month(year, month)

    print(f"  יום 1 בעמודה {get_column_letter(first_col)}")

    # ניקוי השורות (עמודות F עד AO)
    for col_num in range(6, 42):
        col_letter = get_column_letter(col_num)
        # ניקוי תאריכים לועזיים
        ws[f'{col_letter}{greg_row}'] = None
        # ניקוי תאריכים עבריים
        ws[f'{col_letter}{heb_row}'] = None
        # ניקוי אירועים (3 שורות)
        for event_offset in range(3):
            ws[f'{col_letter}{events_row + event_offset}'] = None

    # מילוי התאריכים
    for day in range(1, num_days + 1):
        current_date = date(year, month, day)
        col_num = first_col + day - 1
        col_letter = get_column_letter(col_num)
        weekday = current_date.weekday()

        # קביעת צבע לפי יום בשבוע
        if weekday == 5:  # Saturday - שבת
            fill = PatternFill(start_color=SHABBAT_COLOR, end_color=SHABBAT_COLOR, fill_type='solid')
        elif weekday == 4:  # Friday - שישי
            fill = PatternFill(start_color=FRIDAY_COLOR, end_color=FRIDAY_COLOR, fill_type='solid')
        else:
            fill = PatternFill(start_color=WHITE_COLOR, end_color=WHITE_COLOR, fill_type='solid')

        # תאריך לועזי
        cell = ws[f'{col_letter}{greg_row}']
        cell.value = day
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = fill

        # תאריך עברי
        cell = ws[f'{col_letter}{heb_row}']
        cell.value = get_hebrew_date_str(current_date)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = fill

        # אירועים
        date_key = (year, month, day)
        event_text = None

        if date_key in JEWISH_HOLIDAYS:
            event_text = JEWISH_HOLIDAYS[date_key]
        elif date_key in ARAB_HOLIDAYS:
            event_text = ARAB_HOLIDAYS[date_key]
        elif date_key in DRUZE_HOLIDAYS:
            event_text = DRUZE_HOLIDAYS[date_key]

        cell = ws[f'{col_letter}{events_row}']
        cell.value = event_text
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.fill = fill

# הוספת מסגרות לכל התאים בטווח
print("\nמוסיף מסגרות...")
for row in range(5, 70):
    for col_num in range(6, 42):
        col_letter = get_column_letter(col_num)
        cell = ws[f'{col_letter}{row}']
        if not is_merged_cell(ws, f'{col_letter}{row}'):
            cell.border = thin_border

# שמירה
wb.save(output_path)
print(f'\nהקובץ נשמר: {output_path}')
