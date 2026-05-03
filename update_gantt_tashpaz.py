# -*- coding: utf-8 -*-
"""
עדכון קובץ גאנט מתשפ"ו לתשפ"ז
שומר על אותו עיצוב, רק מעדכן תאריכים וחגים
"""
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.cell_range import CellRange
from pyluach import dates
from datetime import date, timedelta
import calendar
import copy

# קובץ מקור
source_path = r'C:\Users\Shahar Tabib\OneDrive - Sliceknowledge.com\Slice\שתפים\מישרים\רות זהר גאנט תשפז\לוח שנתי מישרים תשפז.xlsx'
output_path = r'C:\Users\Shahar Tabib\OneDrive - Sliceknowledge.com\Slice\שתפים\מישרים\רות זהר גאנט תשפז\לוח שנתי מישרים תשפז - חדש.xlsx'

# טעינת הקובץ
wb = openpyxl.load_workbook(source_path)
ws = wb['גאנט תשפו']

# שינוי שם הגיליון
ws.title = "גאנט תשפז"

# בדיקה אם תא הוא חלק מתאים ממוזגים
def is_merged_cell(ws, row, col):
    cell_coord = f'{get_column_letter(col)}{row}'
    for merged_range in ws.merged_cells.ranges:
        if cell_coord in merged_range:
            return True
    return False

# מבנה החודשים בקובץ המקורי:
# שורה 5: ימים בשבוע (ד', ה', ו', ש', א', ב', ג'...) - זה לא ישתנה
# שורה 6: תאריכים לועזיים ספטמבר
# שורה 7: תאריכים עבריים ספטמבר
# שורות 8-10: אירועים
# שורה 11: תאריכים לועזיים אוקטובר
# ...

# חגים יהודיים תשפ"ז (5787)
JEWISH_HOLIDAYS = {
    (2026, 9, 11): "ערה\"ש",
    (2026, 9, 12): "ראש השנה",
    (2026, 9, 13): "ראש השנה",
    (2026, 9, 14): "צום גדליה",
    (2026, 9, 20): "עיו\"כ",
    (2026, 9, 21): "יום כיפור",
    (2026, 9, 22): "חופש",
    (2026, 9, 25): "ערב סוכות",
    (2026, 9, 26): "סוכות",
    (2026, 9, 27): "חוה\"מ סוכות",
    (2026, 9, 28): "חוה\"מ סוכות",
    (2026, 9, 29): "חוה\"מ סוכות",
    (2026, 9, 30): "חוה\"מ סוכות",
    (2026, 10, 1): "חוה\"מ",
    (2026, 10, 2): "הושענא רבה",
    (2026, 10, 3): "שמחת תורה",
    (2026, 10, 7): "יוה\"ז ל-7.10",
    (2026, 11, 4): "יוה\"ז לי.רבין",
    (2026, 12, 6): "חנוכה",
    (2026, 12, 7): "חנוכה",
    (2026, 12, 8): "חנוכה",
    (2026, 12, 9): "חנוכה",
    (2026, 12, 10): "חנוכה",
    (2026, 12, 11): "חנוכה",
    (2026, 12, 12): "חנוכה",
    (2026, 12, 13): "חנוכה",
    (2026, 12, 24): "צום עשרה בטבת",
    (2027, 2, 6): "ט\"ו בשבט",
    (2027, 2, 25): "ת. אסתר",
    (2027, 2, 28): "פורים",
    (2027, 3, 1): "התחלת הרמדאן",
    (2027, 3, 16): "שאלון מורים",
    (2027, 3, 29): "עיד אל פיטר",
    (2027, 3, 31): "חופשת פסח בבתי ספר",
    (2027, 4, 2): "ערב פסח",
    (2027, 4, 3): "פסח",
    (2027, 4, 4): "חוה\"מ פסח",
    (2027, 4, 5): "חוה\"מ פסח",
    (2027, 4, 6): "חוה\"מ פסח",
    (2027, 4, 7): "חוה\"מ פסח",
    (2027, 4, 8): "חוה\"מ פסח",
    (2027, 4, 9): "שביעי של פסח",
    (2027, 4, 10): "אסרו חג",
    (2027, 4, 19): "מחוון אוריינות",
    (2027, 4, 21): "יוה\"ז לחללי מערכות ישראל",
    (2027, 4, 22): "יום העצמאות",
    (2027, 4, 25): "חג הנביא שועייב",
    (2027, 5, 5): "ל\"ג בעומר",
    (2027, 5, 15): "יום ירושלים",
    (2027, 5, 22): "ערב שבועות",
    (2027, 5, 23): "שבועות",
    (2027, 6, 7): "עיד אל אדחא",
    (2027, 7, 4): "צום י\"ז בתמוז",
    (2027, 7, 25): "צום ט' באב",
    (2027, 6, 30): "משימות סוף שנה כיתות א",
}

# חגים ערביים
ARAB_HOLIDAYS = {
    (2027, 3, 1): "התחלת הרמדאן",
    (2027, 3, 29): "עיד אל פיטר",
    (2027, 6, 7): "עיד אל אדחא",
}

# חגים דרוזיים
DRUZE_HOLIDAYS = {
    (2027, 4, 25): "חג הנביא שועייב",
}

# חופשות
VACATIONS = {
    "חופשת אביב חברה ערבית": [(2027, 4, 2)],
    "חופשת אביב חברה דרוזית": [(2027, 4, 1), (2027, 3, 31)],
}


def get_hebrew_date_str(d):
    """מחזיר תאריך עברי בפורמט קריא"""
    try:
        heb = dates.HebrewDate.from_pydate(d)
        day = heb.day
        hebrew_nums = {
            1: "א'", 2: "ב'", 3: "ג'", 4: "ד'", 5: "ה'", 6: "ו'", 7: "ז'", 8: "ח'", 9: "ט'",
            10: "י'", 11: 'י"א', 12: 'י"ב', 13: 'י"ג', 14: 'י"ד', 15: 'ט"ו', 16: 'ט"ז',
            17: 'י"ז', 18: 'י"ח', 19: 'י"ט', 20: "כ'", 21: 'כ"א', 22: 'כ"ב', 23: 'כ"ג',
            24: 'כ"ד', 25: 'כ"ה', 26: 'כ"ו', 27: 'כ"ז', 28: 'כ"ח', 29: 'כ"ט', 30: "ל'"
        }
        return hebrew_nums.get(day, str(day))
    except:
        return ""


def get_weekday_hebrew(d):
    """מחזיר יום בשבוע בעברית"""
    # Monday=0, ..., Sunday=6
    # בעברית: א'=ראשון(Sunday=6), ב'=שני(Monday=0), וכו'
    days = ["ב'", "ג'", "ד'", "ה'", "ו'", "ש'", "א'"]
    return days[d.weekday()]


# מבנה החודשים - שורות תאריכים לועזיים
# בקובץ המקורי: ספטמבר בשורה 6, אוקטובר בשורה 11, וכו'
MONTH_ROWS = {
    (2024, 9): 6,   # ספטמבר 2024 בקובץ המקורי
    (2024, 10): 11, # אוקטובר
    (2024, 11): 16, # נובמבר
    (2024, 12): 21, # דצמבר
    (2025, 1): 26,  # ינואר
    (2025, 2): 31,  # פברואר
    (2025, 3): 36,  # מרץ
    (2025, 4): 41,  # אפריל
    (2025, 5): 46,  # מאי
    (2025, 6): 52,  # יוני
    (2025, 7): 57,  # יולי
    (2025, 8): 62,  # אוגוסט
}

# מיפוי לשנה החדשה תשפ"ז
NEW_MONTH_ROWS = {
    (2026, 9): 6,   # ספטמבר 2026
    (2026, 10): 11, # אוקטובר
    (2026, 11): 16, # נובמבר
    (2026, 12): 21, # דצמבר
    (2027, 1): 26,  # ינואר
    (2027, 2): 31,  # פברואר
    (2027, 3): 36,  # מרץ
    (2027, 4): 41,  # אפריל
    (2027, 5): 46,  # מאי
    (2027, 6): 52,  # יוני
    (2027, 7): 57,  # יולי
    (2027, 8): 62,  # אוגוסט
}

# מציאת העמודה של יום 1 לכל חודש בקובץ המקור
FIRST_DAY_COLS = {
    6: 'I',   # ספטמבר - 1 בעמודה I
    11: 'L',  # אוקטובר
    16: 'G',  # נובמבר
    21: 'J',  # דצמבר
    26: 'L',  # ינואר
    31: 'H',  # פברואר
    36: 'K',  # מרץ
    41: 'J',  # אפריל
    46: 'F',  # מאי
    52: 'H',  # יוני
    57: 'K',  # יולי
    62: 'F',  # אוגוסט
}


def col_to_num(col_letter):
    """המרת אות עמודה למספר"""
    result = 0
    for char in col_letter:
        result = result * 26 + (ord(char.upper()) - ord('A') + 1)
    return result


def clear_events_rows(ws, greg_row):
    """מנקה שורות אירועים מתחת לשורות התאריכים"""
    # שורות אירועים הן בדרך כלל 2-4 שורות מתחת לשורת התאריך העברי
    heb_row = greg_row + 1
    for event_row in range(heb_row + 1, heb_row + 4):
        for col in range(1, 43):
            letter = get_column_letter(col)
            cell = ws[f'{letter}{event_row}']
            # שומר על הצבע אבל מנקה את הטקסט
            # לא מוחק - כי אנחנו רוצים לשמור על העיצוב


# עדכון התאריכים והחגים
for (year, month), greg_row in NEW_MONTH_ROWS.items():
    heb_row = greg_row + 1
    events_row = greg_row + 2  # שורת אירועים ראשית

    # מספר ימים בחודש
    num_days = calendar.monthrange(year, month)[1]

    # מציאת היום הראשון בחודש
    first_day = date(year, month, 1)
    first_weekday = first_day.weekday()  # 0=Monday, 6=Sunday

    # בקובץ המקורי, העמודות מתחילות מ-F (יום ד')
    # שורה 5 מכילה: F=ד', G=ה', H=ו', I=ש', J=א', K=ב', L=ג', M=ד'...
    # המיפוי: ד'=3(Wed), ה'=4(Thu), ו'=5(Fri), ש'=6(Sat), א'=0(Sun), ב'=1(Mon), ג'=2(Tue)

    # מיפוי יום בשבוע לאינדקס עמודה (מתחיל מעמודה F=6)
    # עמודה F=6 היא יום ד' (weekday 2 in Python where Mon=0)
    # בפייתון: Mon=0, Tue=1, Wed=2, Thu=3, Fri=4, Sat=5, Sun=6
    # העמודות: F(6)=ד'(Wed=2), G(7)=ה'(Thu=3), H(8)=ו'(Fri=4), I(9)=ש'(Sat=5),
    #          J(10)=א'(Sun=6), K(11)=ב'(Mon=0), L(12)=ג'(Tue=1), M(13)=ד'(Wed=2)...

    # חישוב העמודה הראשונה
    # F=ד'(Wed=2), cycle of 7 columns per week
    base_col = 6  # עמודה F
    base_weekday = 2  # Wednesday

    # חישוב ההפרש
    if first_weekday >= base_weekday:
        offset = first_weekday - base_weekday
    else:
        offset = 7 - (base_weekday - first_weekday)

    start_col = base_col + offset

    print(f"\n{year}/{month}: first_weekday={first_weekday}, offset={offset}, start_col={start_col} ({get_column_letter(start_col)})")

    # עדכון התאריכים
    for day in range(1, num_days + 1):
        col_num = start_col + day - 1
        col_letter = get_column_letter(col_num)
        current_date = date(year, month, day)

        # עדכון תאריך לועזי
        ws[f'{col_letter}{greg_row}'] = day

        # עדכון תאריך עברי
        ws[f'{col_letter}{heb_row}'] = get_hebrew_date_str(current_date)

        # בדיקת חגים ואירועים - רק אם התא לא ממוזג
        date_key = (year, month, day)
        col_num_for_check = col_num
        if date_key in JEWISH_HOLIDAYS:
            if not is_merged_cell(ws, events_row, col_num_for_check):
                ws[f'{col_letter}{events_row}'] = JEWISH_HOLIDAYS[date_key]
            else:
                print(f'  Skipping merged cell: {col_letter}{events_row}')

# שמירת הקובץ
wb.save(output_path)
print(f'\nהקובץ נשמר בהצלחה: {output_path}')
